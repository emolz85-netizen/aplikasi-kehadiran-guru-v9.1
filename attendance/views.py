import base64
import io
import math
import calendar
import json
import hashlib
import random
from datetime import time, datetime, timedelta
from PIL import Image, ImageStat

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.contrib.auth.hashers import make_password, check_password
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import connection, models
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.urls import reverse

from .forms import LeaveRequestForm, OfficialDutyForm, TeacherImportForm, ProfileForm, MalayPasswordChangeForm, PasswordRecoveryRequestForm, PasswordRecoveryConfirmForm, QRPasswordSetForm, LeaveReviewForm, FaceReferenceForm
from .models import Attendance, LeaveRequest, OfficialDuty, TeacherProfile, SchoolSettings, SchoolHoliday, AccountActivity, PasswordRecoveryRequest, PushSubscription, AppNotification, TrustedDevice, LocationSecurityEvent

LIVENESS_CHALLENGES = [
    "Senyum dan pandang terus ke kamera",
    "Pusing kepala sedikit ke kiri",
    "Pusing kepala sedikit ke kanan",
    "Angkat kening dan pandang kamera",
]

def _image_fingerprint(upload):
    upload.seek(0)
    raw = upload.read()
    upload.seek(0)
    digest = hashlib.sha256(raw).hexdigest()
    img = Image.open(io.BytesIO(raw)).convert("L")
    width, height = img.size
    stat = ImageStat.Stat(img)
    brightness = stat.mean[0]
    contrast = stat.stddev[0]
    small = img.resize((16, 16))
    pixels = list(small.getdata())
    avg = sum(pixels) / len(pixels)
    bits = [1 if px >= avg else 0 for px in pixels]
    return digest, bits, width, height, brightness, contrast

def _visual_match(reference, selfie):
    _, ref_bits, _, _, _, _ = _image_fingerprint(reference)
    digest, live_bits, width, height, brightness, contrast = _image_fingerprint(selfie)
    same = sum(a == b for a, b in zip(ref_bits, live_bits))
    score = round((same / len(ref_bits)) * 100, 1)
    quality_ok = width >= 320 and height >= 240 and 35 <= brightness <= 220 and contrast >= 18
    return digest, score, quality_ok, {"width": width, "height": height, "brightness": round(brightness,1), "contrast": round(contrast,1)}



def _risk_level(score):
    if score >= 61: return "TINGGI"
    if score >= 21: return "SEDERHANA"
    return "RENDAH"

def _parse_float(value, default=None):
    try: return float(value)
    except (TypeError, ValueError): return default

def _evaluate_location_security(request, school, user, lat, lng, accuracy, device_id):
    flags, score = [], 0
    now_ms = timezone.now().timestamp() * 1000
    location_ts = _parse_float(request.POST.get("location_timestamp"), 0)
    age_seconds = max(0, (now_ms - location_ts) / 1000) if location_ts else 9999
    if age_seconds > school.max_location_age_seconds:
        flags.append(f"Bacaan GPS lama ({age_seconds:.0f}s)"); score += 25
    if accuracy > school.max_gps_accuracy_meters:
        flags.append(f"Ketepatan lemah ({accuracy:.0f}m)"); score += 25
    if accuracy <= 1:
        flags.append("Ketepatan luar biasa sempurna"); score += 15
    if not device_id or len(device_id) != 64:
        flags.append("ID peranti tidak sah"); score += 30

    device = None
    if school.device_trust_enabled and device_id:
        trusted_exists = TrustedDevice.objects.filter(user=user, status="DIPERCAYAI").exists()
        default_status = "DIPERCAYAI" if school.auto_trust_first_device and not trusted_exists else "MENUNGGU"
        device, created = TrustedDevice.objects.get_or_create(user=user, device_id=device_id, defaults={
            "device_name": (request.POST.get("device_name") or "Peranti web")[:160],
            "platform": (request.POST.get("device_platform") or "")[:100],
            "browser": (request.POST.get("device_browser") or "")[:100],
            "user_agent": request.META.get("HTTP_USER_AGENT", "")[:2000],
            "status": default_status, "first_ip": get_client_ip(request), "last_ip": get_client_ip(request),
            "approved_at": timezone.now() if default_status == "DIPERCAYAI" else None,
        })
        if not created:
            device.last_ip = get_client_ip(request); device.device_name=(request.POST.get("device_name") or device.device_name)[:160]; device.save()
        if device.status == "DISEKAT": flags.append("Peranti disekat"); score += 70
        elif device.status != "DIPERCAYAI": flags.append("Peranti belum diluluskan"); score += 35

    previous = Attendance.objects.filter(user=user).exclude(check_in_lat__isnull=True).order_by("-check_in").first()
    speed_kmh = None
    if previous and previous.check_in and previous.check_in_lat is not None:
        elapsed_h = max((timezone.now() - previous.check_in).total_seconds() / 3600, 0.001)
        moved_km = haversine_m(lat, lng, previous.check_in_lat, previous.check_in_lng) / 1000
        speed_kmh = moved_km / elapsed_h
        if speed_kmh > school.max_plausible_speed_kmh:
            flags.append(f"Pergerakan tidak munasabah ({speed_kmh:.0f} km/j)"); score += 45
    score = min(score, 100)
    return {"score": score, "level": _risk_level(score), "flags": flags, "device": device, "age_seconds": age_seconds, "speed_kmh": speed_kmh}

def health(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        return JsonResponse({"status": "ok", "database": "connected", "version": "9.2.007.1"})
    except Exception:
        return JsonResponse({"status": "error", "database": "unavailable", "version": "9.2.007.1"}, status=503)

def manifest(request):
    return JsonResponse({
        "id": "/?source=pwa",
        "name": f"Sistem Kehadiran Guru {settings.SCHOOL_NAME}",
        "short_name": "Kehadiran",
        "description": f"Aplikasi kehadiran guru {settings.SCHOOL_NAME} dengan GPS, swafoto, cuti dan notifikasi.",
        "lang": "ms-MY", "dir": "ltr", "start_url": "/?source=pwa", "scope": "/",
        "display": "standalone", "display_override": ["standalone", "minimal-ui"],
        "orientation": "portrait-primary", "background_color": "#f5f6f7", "theme_color": "#1f2937",
        "categories": ["productivity", "education", "utilities"],
        "icons": [
            {"src": "/static/attendance/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
            {"src": "/static/attendance/icon-maskable-192.png", "sizes": "192x192", "type": "image/png", "purpose": "maskable"},
            {"src": "/static/attendance/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
            {"src": "/static/attendance/icon-maskable-512.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"}
        ],
        "shortcuts": [
            {"name": "Rekod Kehadiran", "short_name": "Kehadiran", "url": "/?source=shortcut", "icons": [{"src": "/static/attendance/icon-192.png", "sizes": "192x192"}]},
            {"name": "Permohonan Cuti", "short_name": "Cuti", "url": "/cuti/?source=shortcut", "icons": [{"src": "/static/attendance/icon-192.png", "sizes": "192x192"}]},
            {"name": "Notifikasi", "short_name": "Notifikasi", "url": "/notifikasi/?source=shortcut", "icons": [{"src": "/static/attendance/icon-192.png", "sizes": "192x192"}]}
        ],
        "prefer_related_applications": False
    }, json_dumps_params={"ensure_ascii": False})


def offline_page(request):
    return render(request, "attendance/offline.html")


def pwa_install(request):
    return render(request, "attendance/pwa_install.html")


def service_worker(request):
    script = r'''const VERSION = "9.2.007.1";
const STATIC_CACHE = `kehadiran-static-${VERSION}`;
const PAGE_CACHE = `kehadiran-pages-${VERSION}`;
const OFFLINE_URL = "/offline/";
const PRECACHE = [OFFLINE_URL,"/login/","/pwa/pasang/","/manifest.json","/static/attendance/style.css","/static/attendance/app.js","/static/attendance/pwa.js","/static/attendance/push.js","/static/attendance/icon-192.png","/static/attendance/icon-512.png","/static/attendance/icon-maskable-192.png","/static/attendance/icon-maskable-512.png","/static/attendance/apple-touch-icon.png"];
self.addEventListener("install",e=>e.waitUntil(caches.open(STATIC_CACHE).then(c=>c.addAll(PRECACHE)).then(()=>self.skipWaiting())));
self.addEventListener("activate",e=>e.waitUntil(caches.keys().then(keys=>Promise.all(keys.filter(k=>![STATIC_CACHE,PAGE_CACHE].includes(k)).map(k=>caches.delete(k)))).then(()=>self.clients.claim())));
self.addEventListener("message",e=>{if(e.data&&e.data.type==="SKIP_WAITING")self.skipWaiting();});
self.addEventListener("fetch",e=>{const r=e.request;if(r.method!=="GET")return;const u=new URL(r.url);if(u.origin!==self.location.origin)return;if(r.mode==="navigate"){e.respondWith(fetch(r).catch(async()=>await caches.match(OFFLINE_URL)));return;}if(u.pathname.startsWith("/static/")){e.respondWith(caches.match(r).then(cached=>{const update=fetch(r).then(resp=>{if(resp.ok)caches.open(STATIC_CACHE).then(c=>c.put(r,resp.clone()));return resp;}).catch(()=>cached);return cached||update;}));}});
self.addEventListener("push",e=>{let d={title:"Sistem Kehadiran Guru",body:"Anda menerima notifikasi baharu.",url:"/notifikasi/"};try{d=Object.assign(d,e.data.json());}catch(x){}e.waitUntil(self.registration.showNotification(d.title,{body:d.body,icon:"/static/attendance/icon-192.png",badge:"/static/attendance/icon-192.png",data:{url:d.url||"/notifikasi/"},tag:"kehadiran-"+(d.notification_id||Date.now()),renotify:true,vibrate:[150,80,150]}));});
self.addEventListener("notificationclick",e=>{e.notification.close();const t=e.notification.data&&e.notification.data.url?e.notification.data.url:"/notifikasi/";e.waitUntil(clients.matchAll({type:"window",includeUncontrolled:true}).then(list=>{for(const c of list){if(c.url.includes(t)&&"focus" in c)return c.focus();}return clients.openWindow?clients.openWindow(t):null;}));});'''
    response = HttpResponse(script, content_type="application/javascript; charset=utf-8")
    response["Service-Worker-Allowed"] = "/"
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response



def get_client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def describe_device(user_agent):
    ua = (user_agent or "").lower()
    if "iphone" in ua:
        platform = "iPhone"
    elif "ipad" in ua:
        platform = "iPad"
    elif "android" in ua:
        platform = "Android"
    elif "windows" in ua:
        platform = "Windows"
    elif "macintosh" in ua or "mac os" in ua:
        platform = "macOS"
    elif "linux" in ua:
        platform = "Linux"
    else:
        platform = "Peranti tidak dikenal pasti"

    if "edg/" in ua:
        browser = "Edge"
    elif "chrome/" in ua and "edg/" not in ua:
        browser = "Chrome"
    elif "firefox/" in ua:
        browser = "Firefox"
    elif "safari/" in ua and "chrome/" not in ua:
        browser = "Safari"
    else:
        browser = "Pelayar tidak dikenal pasti"
    return f"{platform} · {browser}"

def haversine_m(lat1, lon1, lat2, lon2):
    r = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp = math.radians(lat2-lat1)
    dl = math.radians(lon2-lon1)
    a = math.sin(dp/2)**2 + math.cos(p1)*math.cos(p2)*math.sin(dl/2)**2
    return r * (2 * math.atan2(math.sqrt(a), math.sqrt(1-a)))

@login_required
def dashboard(request):
    if request.user.is_staff:
        return redirect("admin_dashboard")
    today = timezone.localdate()
    today_holiday = SchoolHoliday.objects.filter(date=today, is_active=True).first()
    record = Attendance.objects.filter(user=request.user, date=today).first()
    month_records = Attendance.objects.filter(
        user=request.user,
        date__year=today.year,
        date__month=today.month,
    )
    leave_days = set()
    for item in LeaveRequest.objects.filter(
        user=request.user, status="DILULUSKAN",
        start_date__lte=today.replace(day=calendar.monthrange(today.year, today.month)[1]),
        end_date__gte=today.replace(day=1),
    ):
        day = max(item.start_date, today.replace(day=1))
        last = min(item.end_date, today.replace(day=calendar.monthrange(today.year, today.month)[1]))
        while day <= last:
            leave_days.add(day)
            day += timezone.timedelta(days=1)

    duty_days = set()
    for item in OfficialDuty.objects.filter(
        user=request.user, status="DILULUSKAN",
        start_date__lte=today.replace(day=calendar.monthrange(today.year, today.month)[1]),
        end_date__gte=today.replace(day=1),
    ):
        day = max(item.start_date, today.replace(day=1))
        last = min(item.end_date, today.replace(day=calendar.monthrange(today.year, today.month)[1]))
        while day <= last:
            duty_days.add(day)
            day += timezone.timedelta(days=1)

    holiday_days = {item.date: item.name for item in SchoolHoliday.objects.filter(
        is_active=True, date__year=today.year, date__month=today.month
    )}

    records_by_day = {r.date.day: r for r in month_records}
    month_calendar = []
    for week in calendar.Calendar(firstweekday=0).monthdatescalendar(today.year, today.month):
        row = []
        for day in week:
            state = "outside"
            label = ""
            if day.month == today.month:
                state = "empty"
                if day in holiday_days:
                    state, label = "holiday", holiday_days[day]
                elif day in leave_days:
                    state, label = "leave", "Cuti"
                elif day in duty_days:
                    state, label = "duty", "Tugas rasmi"
                elif day.day in records_by_day:
                    rec = records_by_day[day.day]
                    state = "late" if rec.status == "LEWAT" else "present"
                    label = rec.get_status_display()
                elif day.weekday() >= 5:
                    state, label = "weekend", "Hujung minggu"
                elif day < today:
                    state, label = "absent", "Tiada rekod"
            row.append({"date": day, "day": day.day, "state": state, "label": label})
        month_calendar.append(row)

    school = SchoolSettings.load()
    target_in, target_out = school.times_for_date(today)
    week_start = today - timezone.timedelta(days=today.weekday())
    week_records = Attendance.objects.filter(
        user=request.user,
        date__gte=week_start,
        date__lte=today,
        check_in__isnull=False,
    )
    today_status = "Belum daftar masuk"
    today_status_class = "pending"
    if record and record.check_out:
        today_status = "Selesai daftar keluar"
        today_status_class = "complete"
    elif record and record.check_in:
        today_status = "Sudah daftar masuk"
        today_status_class = "active"

    return render(request, "attendance/dashboard.html", {
        "record": record,
        "today": today,
        "month_count": month_records.count(),
        "late_count": month_records.filter(status="LEWAT").count(),
        "school": school,
        "target_in": target_in,
        "target_out": target_out,
        "month_calendar": month_calendar,
        "month_name": today.strftime("%B %Y"),
        "today_holiday": today_holiday,
        "week_count": week_records.count(),
        "today_status": today_status,
        "today_status_class": today_status_class,
        "liveness_challenge": random.choice(LIVENESS_CHALLENGES),
        "face_settings": school,
        "has_reference_photo": TeacherProfile.objects.filter(user=request.user, reference_photo__isnull=False).exclude(reference_photo="").exists(),
    })

@login_required
@require_POST
def record_attendance(request, action):
    if request.user.is_staff:
        return JsonResponse({"ok": False, "message": "Akaun pentadbir tidak dibenarkan merekod kehadiran."}, status=403)
    if action not in {"masuk", "keluar"}:
        return JsonResponse({"ok": False, "message": "Tindakan tidak sah."}, status=400)

    try:
        lat = float(request.POST["latitude"])
        lng = float(request.POST["longitude"])
        accuracy = float(request.POST.get("accuracy", 9999))
    except (KeyError, ValueError):
        return JsonResponse({"ok": False, "message": "Lokasi GPS tidak sah."}, status=400)

    school = SchoolSettings.load()
    if accuracy > school.max_gps_accuracy_meters:
        return JsonResponse({
            "ok": False,
            "message": f"Ketepatan GPS masih lemah ({accuracy:.0f} m). Cuba Segarkan GPS sehingga {school.max_gps_accuracy_meters} m atau lebih baik."
        }, status=400)

    distance = haversine_m(lat, lng, school.latitude, school.longitude)
    if distance > school.radius_meters:
        return JsonResponse({
            "ok": False,
            "message": f"Anda berada {distance:.1f} m dari sekolah. Had ialah {school.radius_meters} m."
        }, status=403)

    today = timezone.localdate()
    holiday = SchoolHoliday.objects.filter(date=today, is_active=True).first()
    if holiday:
        return JsonResponse({
            "ok": False,
            "message": f"Hari ini ialah {holiday.name}. Kehadiran tidak diperlukan."
        }, status=403)
    if today.weekday() >= 5:
        return JsonResponse({
            "ok": False,
            "message": "Hari ini ialah hujung minggu. Kehadiran tidak diperlukan."
        }, status=403)
    now = timezone.now()
    rec, _ = Attendance.objects.get_or_create(user=request.user, date=today)
    selfie = request.FILES.get("selfie")
    user_agent = request.META.get("HTTP_USER_AGENT", "")[:2000]
    device = describe_device(user_agent)
    client_ip = get_client_ip(request)
    device_id = (request.POST.get("device_id") or "").strip()[:64]
    security = _evaluate_location_security(request, school, request.user, lat, lng, accuracy, device_id)
    blocked = security["score"] >= school.high_risk_block_threshold or (security["device"] and security["device"].status == "DISEKAT") or (school.block_untrusted_device and security["device"] and security["device"].status != "DIPERCAYAI")
    if blocked:
        LocationSecurityEvent.objects.create(user=request.user, event_type="RISIKO", action=action, risk_score=security["score"], risk_level=security["level"], flags=security["flags"], latitude=lat, longitude=lng, accuracy=accuracy, device_id=device_id, ip_address=client_ip, blocked=True, details="Rekod disekat oleh kawalan Anti GPS Spoofing & Device Trust")
        return JsonResponse({"ok": False, "message": f"Rekod disekat kerana risiko keselamatan {security['level'].lower()} ({security['score']}/100). Hubungi pentadbir.", "risk_score": security["score"], "risk_level": security["level"], "flags": security["flags"]}, status=403)
    challenge = (request.POST.get("liveness_challenge") or "").strip()[:120]
    liveness_confirmed = request.POST.get("liveness_confirmed") == "1"
    face_score = None
    face_status = "TIDAK_AKTIF"
    selfie_hash = ""

    if school.face_verification_enabled:
        profile, _ = TeacherProfile.objects.get_or_create(user=request.user)
        if not profile.reference_photo:
            return JsonResponse({"ok": False, "message": "Sila muat naik foto rujukan wajah di halaman Profil terlebih dahulu."}, status=400)
        if not selfie:
            return JsonResponse({"ok": False, "message": "Swafoto wajib untuk pengesahan identiti."}, status=400)
        if school.require_liveness_challenge and (not challenge or not liveness_confirmed):
            return JsonResponse({"ok": False, "message": "Sila lengkapkan dan sahkan cabaran hidup yang dipaparkan."}, status=400)
        try:
            selfie_hash, face_score, quality_ok, quality = _visual_match(profile.reference_photo, selfie)
        except Exception:
            return JsonResponse({"ok": False, "message": "Imej tidak dapat diproses. Ambil semula swafoto yang jelas."}, status=400)
        if not quality_ok:
            return JsonResponse({"ok": False, "message": "Kualiti swafoto tidak mencukupi. Pastikan wajah terang, jelas dan kamera stabil."}, status=400)
        duplicate = Attendance.objects.filter(models.Q(selfie_in_hash=selfie_hash) | models.Q(selfie_out_hash=selfie_hash)).exclude(pk=rec.pk).exists()
        if duplicate:
            return JsonResponse({"ok": False, "message": "Imej yang sama pernah digunakan. Ambil swafoto baharu secara langsung."}, status=403)
        face_status = "LULUS" if face_score >= school.face_match_threshold else "SEMAKAN"
        if face_status != "LULUS":
            return JsonResponse({"ok": False, "message": f"Padanan visual hanya {face_score:.1f}%. Ambil semula dengan wajah menghadap kamera dan pencahayaan yang baik."}, status=403)

    if action == "masuk":
        if rec.check_in:
            return JsonResponse({"ok": False, "message": "Rekod masuk sudah dibuat."}, status=400)
        rec.check_in = now
        rec.check_in_lat = lat
        rec.check_in_lng = lng
        rec.check_in_accuracy = accuracy
        rec.distance_in_m = distance
        rec.selfie_in = selfie
        rec.check_in_ip = client_ip
        rec.check_in_device = device
        rec.check_in_user_agent = user_agent
        rec.face_in_score = face_score
        rec.face_in_status = face_status
        rec.liveness_in_challenge = challenge
        rec.selfie_in_hash = selfie_hash
        rec.check_in_device_id = device_id
        rec.check_in_risk_score = security["score"]
        rec.check_in_risk_level = security["level"]
        rec.check_in_security_flags = security["flags"]
        target_in, _ = school.times_for_date(today)
        rec.status = "LEWAT" if timezone.localtime(now).time() > target_in else "HADIR"
    else:
        if not rec.check_in:
            return JsonResponse({"ok": False, "message": "Sila rekod masuk dahulu."}, status=400)
        if rec.check_out:
            return JsonResponse({"ok": False, "message": "Rekod keluar sudah dibuat."}, status=400)
        rec.check_out = now
        rec.check_out_lat = lat
        rec.check_out_lng = lng
        rec.check_out_accuracy = accuracy
        rec.distance_out_m = distance
        rec.selfie_out = selfie
        rec.check_out_ip = client_ip
        rec.check_out_device = device
        rec.check_out_user_agent = user_agent
        rec.face_out_score = face_score
        rec.face_out_status = face_status
        rec.liveness_out_challenge = challenge
        rec.selfie_out_hash = selfie_hash
        rec.check_out_device_id = device_id
        rec.check_out_risk_score = security["score"]
        rec.check_out_risk_level = security["level"]
        rec.check_out_security_flags = security["flags"]

    rec.save()
    LocationSecurityEvent.objects.create(user=request.user, attendance=rec, event_type="RISIKO", action=action, risk_score=security["score"], risk_level=security["level"], flags=security["flags"], latitude=lat, longitude=lng, accuracy=accuracy, device_id=device_id, ip_address=client_ip, blocked=False, details="; ".join(security["flags"]) or "Semakan keselamatan lulus")
    from .push import send_notification
    local_time = timezone.localtime(now).strftime("%H:%M")
    if action == "masuk":
        send_notification(request.user, "Daftar masuk berjaya", f"Kehadiran anda telah direkodkan pada {local_time}.", "KEHADIRAN", "/")
    else:
        send_notification(request.user, "Daftar keluar berjaya", f"Rekod pulang anda telah disimpan pada {local_time}.", "KEHADIRAN", "/")
    return JsonResponse({"ok": True, "message": f"Berjaya. Jarak {distance:.1f} m, ketepatan GPS {accuracy:.0f} m, risiko {security['level'].lower()} ({security['score']}/100).", "risk_score": security["score"], "risk_level": security["level"], "flags": security["flags"]})

@user_passes_test(lambda u: u.is_staff)
def device_security_page(request):
    devices = TrustedDevice.objects.select_related("user", "approved_by").all()
    status = request.GET.get("status", "")
    if status: devices = devices.filter(status=status)
    events = LocationSecurityEvent.objects.select_related("user", "attendance")[:100]
    return render(request, "attendance/device_security.html", {
        "devices": devices, "events": events, "status_filter": status,
        "pending_count": TrustedDevice.objects.filter(status="MENUNGGU").count(),
        "trusted_count": TrustedDevice.objects.filter(status="DIPERCAYAI").count(),
        "blocked_count": TrustedDevice.objects.filter(status="DISEKAT").count(),
        "high_risk_count": LocationSecurityEvent.objects.filter(risk_level="TINGGI").count(),
    })

@require_POST
@user_passes_test(lambda u: u.is_staff)
def device_update_status(request, pk, status):
    if status not in {"DIPERCAYAI", "MENUNGGU", "DISEKAT"}:
        return JsonResponse({"ok": False}, status=400)
    device = get_object_or_404(TrustedDevice, pk=pk)
    device.status = status
    device.approved_by = request.user if status == "DIPERCAYAI" else None
    device.approved_at = timezone.now() if status == "DIPERCAYAI" else None
    device.save()
    messages.success(request, f"Status peranti {device.device_name or device.device_id[:12]} dikemas kini.")
    return redirect("device_security_page")

@login_required
def profile_page(request):
    profile_form = ProfileForm(instance=request.user, prefix="profile")
    password_form = MalayPasswordChangeForm(request.user, prefix="password")
    teacher_profile, _ = TeacherProfile.objects.get_or_create(user=request.user)
    face_form = FaceReferenceForm(instance=teacher_profile, prefix="face")

    if request.method == "POST":
        if "save_profile" in request.POST:
            old_username = request.user.username
            profile_form = ProfileForm(request.POST, instance=request.user, prefix="profile")
            if profile_form.is_valid():
                user = profile_form.save()
                AccountActivity.objects.create(
                    user=user,
                    action="Kemas kini profil",
                    details=f"Nama pengguna: {old_username} → {user.username}",
                )
                messages.success(request, "Profil berjaya dikemas kini.")
                return redirect("profile_page")
        elif "save_face" in request.POST:
            face_form = FaceReferenceForm(request.POST, request.FILES, instance=teacher_profile, prefix="face")
            if face_form.is_valid():
                obj = face_form.save(commit=False)
                obj.reference_photo_updated_at = timezone.now()
                obj.save()
                AccountActivity.objects.create(user=request.user, action="Kemas kini foto rujukan wajah")
                messages.success(request, "Foto rujukan wajah berjaya disimpan.")
                return redirect("profile_page")
        elif "change_password" in request.POST:
            password_form = MalayPasswordChangeForm(request.user, request.POST, prefix="password")
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                AccountActivity.objects.create(user=user, action="Tukar kata laluan")
                messages.success(request, "Kata laluan berjaya ditukar.")
                return redirect("profile_page")

    activities = AccountActivity.objects.filter(user=request.user)[:10]
    return render(request, "attendance/profile.html", {
        "profile_form": profile_form,
        "password_form": password_form,
        "activities": activities,
        "face_form": face_form,
        "teacher_profile": teacher_profile,
    })


@user_passes_test(lambda u: u.is_staff)
def admin_dashboard(request):
    today = timezone.localdate()
    User = get_user_model()
    teachers = User.objects.filter(is_active=True, is_staff=False).order_by("first_name", "username")
    records = Attendance.objects.filter(date=today).select_related("user").order_by("check_in")

    attended_ids = set(records.filter(check_in__isnull=False).values_list("user_id", flat=True))
    leave_ids = set(LeaveRequest.objects.filter(
        status="DILULUSKAN", start_date__lte=today, end_date__gte=today
    ).values_list("user_id", flat=True))
    duty_ids = set(OfficialDuty.objects.filter(
        status="DILULUSKAN", start_date__lte=today, end_date__gte=today
    ).values_list("user_id", flat=True))

    total = teachers.count()
    attended = len(attended_ids)
    late = records.filter(check_in__isnull=False, status="LEWAT").count()
    on_leave = teachers.filter(id__in=leave_ids).count()
    on_duty = teachers.filter(id__in=duty_ids).count()
    excused_ids = leave_ids | duty_ids
    absent = teachers.exclude(id__in=attended_ids | excused_ids)
    percentage = round((attended / total * 100), 1) if total else 0

    # Statistik tujuh hari terakhir untuk graf ringkas tanpa pustaka luaran.
    weekly_stats = []
    max_week_count = 1
    for offset in range(6, -1, -1):
        day = today - timezone.timedelta(days=offset)
        count = Attendance.objects.filter(date=day, check_in__isnull=False).count()
        max_week_count = max(max_week_count, count)
        weekly_stats.append({"date": day, "count": count})
    for item in weekly_stats:
        item["height"] = round((item["count"] / max_week_count) * 100)

    school = SchoolSettings.load()
    today_holiday = SchoolHoliday.objects.filter(date=today, is_active=True).first()
    target_in, target_out = school.times_for_date(today)

    return render(request, "attendance/admin_dashboard.html", {
        "today": today,
        "school": school,
        "target_in": target_in,
        "target_out": target_out,
        "today_holiday": today_holiday,
        "total": total,
        "attended": attended,
        "late": late,
        "on_leave": on_leave,
        "on_duty": on_duty,
        "absent": absent,
        "percentage": percentage,
        "records": records,
        "weekly_stats": weekly_stats,
        "pending_leave": LeaveRequest.objects.filter(status="MENUNGGU").count(),
        "pending_duty": OfficialDuty.objects.filter(status="MENUNGGU").count(),
        "recent_leave_requests": LeaveRequest.objects.select_related("user").filter(status="MENUNGGU")[:5],
    })


@login_required
def leave_page(request):
    if request.user.is_staff:
        messages.info(request, "Akaun pentadbir menggunakan modul Pengurusan Cuti, bukan permohonan cuti peribadi.")
        return redirect("leave_admin")
    if request.method == "POST":
        form = LeaveRequestForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = request.user
            item.save()
            AccountActivity.objects.create(
                user=request.user,
                action="Mohon cuti",
                details=f"{item.get_leave_type_display()}: {item.start_date:%d/%m/%Y} - {item.end_date:%d/%m/%Y}",
            )
            from .push import send_notification
            send_notification(request.user, "Permohonan cuti diterima", "Permohonan cuti anda sedang menunggu semakan pentadbir.", "CUTI", "/cuti/")
            for admin_user in get_user_model().objects.filter(is_staff=True, is_active=True):
                send_notification(admin_user, "Permohonan cuti baharu", f"{request.user.get_full_name() or request.user.username} menghantar permohonan cuti.", "CUTI", "/pengurusan-cuti/")
            messages.success(request, "Permohonan cuti berjaya dihantar kepada pentadbir.")
            return redirect("leave_page")
    else:
        form = LeaveRequestForm(user=request.user)

    items = LeaveRequest.objects.filter(user=request.user)
    current_year = timezone.localdate().year
    approved_crk = sum(
        x.total_days for x in items.filter(
            leave_type="CRK", status="DILULUSKAN", start_date__year=current_year
        )
    )
    return render(request, "attendance/leave.html", {
        "form": form,
        "items": items,
        "pending_count": items.filter(status="MENUNGGU").count(),
        "approved_count": items.filter(status="DILULUSKAN").count(),
        "rejected_count": items.filter(status="DITOLAK").count(),
        "approved_crk_days": approved_crk,
        "current_year": current_year,
    })


@login_required
@require_POST
def leave_cancel(request, pk):
    item = get_object_or_404(LeaveRequest, pk=pk, user=request.user)
    if item.status != "MENUNGGU":
        messages.error(request, "Hanya permohonan yang masih menunggu boleh dibatalkan.")
    else:
        item.status = "DIBATALKAN"
        item.save(update_fields=["status", "updated_at"])
        AccountActivity.objects.create(user=request.user, action="Batal permohonan cuti", details=f"Permohonan #{item.pk}")
        messages.success(request, "Permohonan cuti telah dibatalkan.")
    return redirect("leave_page")


@user_passes_test(lambda u: u.is_staff)
def leave_admin(request):
    status = request.GET.get("status", "MENUNGGU").upper()
    valid_status = {x[0] for x in LeaveRequest.STATUS_CHOICES}
    items = LeaveRequest.objects.select_related("user", "reviewed_by").all()
    if status in valid_status:
        items = items.filter(status=status)
    else:
        status = "SEMUA"
    return render(request, "attendance/leave_admin.html", {
        "items": items,
        "selected_status": status,
        "counts": {
            "MENUNGGU": LeaveRequest.objects.filter(status="MENUNGGU").count(),
            "DILULUSKAN": LeaveRequest.objects.filter(status="DILULUSKAN").count(),
            "DITOLAK": LeaveRequest.objects.filter(status="DITOLAK").count(),
        },
    })


@user_passes_test(lambda u: u.is_staff)
def leave_review(request, pk, decision):
    item = get_object_or_404(LeaveRequest.objects.select_related("user"), pk=pk)
    if item.status != "MENUNGGU":
        messages.warning(request, "Permohonan ini sudah diproses.")
        return redirect("leave_admin")
    form = LeaveReviewForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        if decision not in {"lulus", "tolak"}:
            messages.error(request, "Tindakan tidak sah.")
            return redirect("leave_admin")
        item.status = "DILULUSKAN" if decision == "lulus" else "DITOLAK"
        item.admin_note = form.cleaned_data["admin_note"]
        item.reviewed_by = request.user
        item.reviewed_at = timezone.now()
        item.save(update_fields=["status", "admin_note", "reviewed_by", "reviewed_at", "updated_at"])
        AccountActivity.objects.create(
            user=item.user,
            action="Permohonan cuti " + ("diluluskan" if decision == "lulus" else "ditolak"),
            details=f"Oleh {request.user.get_full_name() or request.user.username}; permohonan #{item.pk}",
        )
        from .push import send_notification
        keputusan = "diluluskan" if decision == "lulus" else "ditolak"
        send_notification(item.user, f"Permohonan cuti {keputusan}", f"Permohonan cuti {item.start_date:%d/%m/%Y} hingga {item.end_date:%d/%m/%Y} telah {keputusan}.", "CUTI", "/cuti/")
        messages.success(request, f"Permohonan {item.user.get_full_name() or item.user.username} telah {item.get_status_display().lower()}.")
        return redirect("leave_admin")
    return render(request, "attendance/leave_review.html", {"item": item, "form": form, "decision": decision})

@login_required
def duty_page(request):
    if request.method == "POST":
        form = OfficialDutyForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.user = request.user
            item.save()
            messages.success(request, "Permohonan tugas rasmi dihantar.")
            return redirect("duty_page")
    else:
        form = OfficialDutyForm()
    return render(request, "attendance/duty.html", {
        "form": form,
        "items": OfficialDuty.objects.filter(user=request.user),
    })

@login_required
def report_page(request):
    today = timezone.localdate()
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(request.GET.get("month", today.month))
        if month < 1 or month > 12:
            raise ValueError
    except (TypeError, ValueError):
        month = today.month

    selected_user_id = request.GET.get("teacher", "").strip()
    status_filter = request.GET.get("status", "").strip().upper()
    report_type = request.GET.get("type", "monthly").strip().lower()
    if report_type not in {"daily", "monthly", "annual", "leave"}:
        report_type = "monthly"

    User = get_user_model()
    teachers = User.objects.filter(is_active=True, is_staff=False).order_by("first_name", "last_name", "username")
    if request.user.is_staff:
        selected_users = teachers
        if selected_user_id.isdigit():
            selected_users = selected_users.filter(pk=int(selected_user_id))
    else:
        selected_users = User.objects.filter(pk=request.user.pk)
        selected_user_id = str(request.user.pk)

    selected_ids = list(selected_users.values_list("id", flat=True))
    records = Attendance.objects.filter(user_id__in=selected_ids, date__year=year).select_related("user")
    if report_type in {"daily", "monthly"}:
        records = records.filter(date__month=month)
    if status_filter in {"HADIR", "LEWAT"}:
        records = records.filter(status=status_filter)
    records = records.order_by("-date", "user__first_name", "user__username")

    leaves = LeaveRequest.objects.filter(
        user_id__in=selected_ids,
        status="DILULUSKAN",
        start_date__year__lte=year,
        end_date__year__gte=year,
    ).select_related("user")
    if report_type in {"daily", "monthly", "leave"}:
        first_day = datetime(year, month, 1).date()
        last_day = datetime(year, month, calendar.monthrange(year, month)[1]).date()
        leaves = leaves.filter(start_date__lte=last_day, end_date__gte=first_day)

    school = SchoolSettings.load()
    month_names = [
        "", "Januari", "Februari", "Mac", "April", "Mei", "Jun",
        "Julai", "Ogos", "September", "Oktober", "November", "Disember"
    ]

    # Ringkasan setiap guru bagi bulan/tahun dipilih.
    summary_rows = []
    period_start = datetime(year, 1, 1).date() if report_type == "annual" else datetime(year, month, 1).date()
    period_end = datetime(year, 12, 31).date() if report_type == "annual" else datetime(year, month, calendar.monthrange(year, month)[1]).date()
    period_end = min(period_end, today) if year == today.year else period_end

    holiday_dates = set(SchoolHoliday.objects.filter(
        is_active=True, date__gte=period_start, date__lte=period_end
    ).values_list("date", flat=True))
    working_days = 0
    cursor = period_start
    while cursor <= period_end:
        if cursor.weekday() < 5 and cursor not in holiday_dates:
            working_days += 1
        cursor += timedelta(days=1)

    for teacher in selected_users:
        teacher_records = records.filter(user=teacher)
        present = teacher_records.filter(check_in__isnull=False).count()
        late = teacher_records.filter(check_in__isnull=False, status="LEWAT").count()
        approved_leave_days = set()
        for leave in LeaveRequest.objects.filter(
            user=teacher, status="DILULUSKAN",
            start_date__lte=period_end, end_date__gte=period_start,
        ):
            day = max(leave.start_date, period_start)
            last = min(leave.end_date, period_end)
            while day <= last:
                if day.weekday() < 5 and day not in holiday_dates:
                    approved_leave_days.add(day)
                day += timedelta(days=1)
        duty_days = set()
        for duty in OfficialDuty.objects.filter(
            user=teacher, status="DILULUSKAN",
            start_date__lte=period_end, end_date__gte=period_start,
        ):
            day = max(duty.start_date, period_start)
            last = min(duty.end_date, period_end)
            while day <= last:
                if day.weekday() < 5 and day not in holiday_dates:
                    duty_days.add(day)
                day += timedelta(days=1)
        absent = max(working_days - present - len(approved_leave_days) - len(duty_days), 0)
        attendance_pct = round((present / working_days * 100), 1) if working_days else 0
        summary_rows.append({
            "teacher": teacher,
            "present": present,
            "late": late,
            "leave": len(approved_leave_days),
            "duty": len(duty_days),
            "absent": absent,
            "percentage": attendance_pct,
        })
    summary_rows.sort(key=lambda row: (-row["percentage"], row["teacher"].get_full_name() or row["teacher"].username))

    total_present = sum(row["present"] for row in summary_rows)
    total_late = sum(row["late"] for row in summary_rows)
    total_leave = sum(row["leave"] for row in summary_rows)
    total_absent = sum(row["absent"] for row in summary_rows)
    average_pct = round(sum(row["percentage"] for row in summary_rows) / len(summary_rows), 1) if summary_rows else 0

    leave_summary = []
    for teacher in selected_users:
        counts = {code: 0 for code, _ in LeaveRequest.LEAVE_TYPE_CHOICES}
        teacher_leaves = leaves.filter(user=teacher)
        for item in teacher_leaves:
            overlap_start = max(item.start_date, period_start)
            overlap_end = min(item.end_date, period_end)
            counts[item.leave_type] += max((overlap_end - overlap_start).days + 1, 0)
        leave_summary.append({"teacher": teacher, "counts": counts, "total": sum(counts.values())})

    query_string = request.GET.urlencode()
    return render(request, "attendance/report.html", {
        "school": school,
        "records": records[:500],
        "teachers": teachers,
        "selected_teacher": selected_user_id,
        "selected_year": year,
        "selected_month": month,
        "selected_status": status_filter,
        "report_type": report_type,
        "month_name": month_names[month],
        "years": range(today.year - 4, today.year + 2),
        "months": [(i, month_names[i]) for i in range(1, 13)],
        "summary_rows": summary_rows,
        "leave_summary": leave_summary,
        "leave_types": LeaveRequest.LEAVE_TYPE_CHOICES,
        "working_days": working_days,
        "total_present": total_present,
        "total_late": total_late,
        "total_leave": total_leave,
        "total_absent": total_absent,
        "average_pct": average_pct,
        "query_string": query_string,
        "generated_at": timezone.localtime(),
    })


def _report_filters(request):
    today = timezone.localdate()
    try:
        year = int(request.GET.get("year", today.year))
    except (TypeError, ValueError):
        year = today.year
    try:
        month = int(request.GET.get("month", today.month))
        if not 1 <= month <= 12:
            raise ValueError
    except (TypeError, ValueError):
        month = today.month
    teacher_id = request.GET.get("teacher", "").strip()
    status = request.GET.get("status", "").strip().upper()
    report_type = request.GET.get("type", "monthly").strip().lower()
    User = get_user_model()
    users = User.objects.filter(is_active=True, is_staff=False)
    if request.user.is_staff:
        if teacher_id.isdigit():
            users = users.filter(pk=int(teacher_id))
    else:
        users = User.objects.filter(pk=request.user.pk)
    records = Attendance.objects.filter(user__in=users, date__year=year).select_related("user")
    if report_type != "annual":
        records = records.filter(date__month=month)
    if status in {"HADIR", "LEWAT"}:
        records = records.filter(status=status)
    return year, month, report_type, users.order_by("first_name", "username"), records.order_by("date", "user__first_name", "user__username")


@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    year, month, report_type, users, records = _report_filters(request)
    school = SchoolSettings.load()
    wb = Workbook()
    ws = wb.active
    ws.title = "Kehadiran"
    ws.merge_cells("A1:H1")
    ws["A1"] = school.school_name
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Laporan Kehadiran {month:02d}/{year}" if report_type != "annual" else f"Laporan Kehadiran Tahunan {year}"
    ws["A2"].font = Font(size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center")
    headers = ["Bil", "Tarikh", "Nama Guru", "Masuk", "Keluar", "Status", "Lewat (minit)", "Catatan"]
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="B8C2CC")
    for cell in ws[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        cell.alignment = Alignment(horizontal="center")
    for idx, r in enumerate(records, 1):
        target_in, _ = school.times_for_date(r.date)
        late_minutes = ""
        if r.check_in and r.status == "LEWAT":
            local_in = timezone.localtime(r.check_in)
            target_dt = timezone.make_aware(datetime.combine(r.date, target_in), timezone.get_current_timezone())
            late_minutes = max(int((local_in - target_dt).total_seconds() // 60), 0)
        ws.append([
            idx, r.date.strftime("%d/%m/%Y"), r.user.get_full_name() or r.user.username,
            timezone.localtime(r.check_in).strftime("%H:%M:%S") if r.check_in else "",
            timezone.localtime(r.check_out).strftime("%H:%M:%S") if r.check_out else "",
            r.get_status_display(), late_minutes, "",
        ])
    ws.auto_filter.ref = f"A4:H{max(ws.max_row, 4)}"
    ws.freeze_panes = "A5"
    widths = [7, 14, 30, 13, 13, 13, 15, 24]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Helaian ringkasan guru.
    summary = wb.create_sheet("Ringkasan")
    summary.append([school.school_name])
    summary.append([f"Ringkasan {month:02d}/{year}" if report_type != "annual" else f"Ringkasan Tahunan {year}"])
    summary.append([])
    summary.append(["Bil", "Nama Guru", "Hadir", "Lewat", "Cuti Diluluskan"])
    for idx, user in enumerate(users, 1):
        user_records = records.filter(user=user)
        leave_qs = LeaveRequest.objects.filter(user=user, status="DILULUSKAN")
        if report_type == "annual":
            leave_qs = leave_qs.filter(start_date__year__lte=year, end_date__year__gte=year)
        else:
            first = datetime(year, month, 1).date()
            last = datetime(year, month, calendar.monthrange(year, month)[1]).date()
            leave_qs = leave_qs.filter(start_date__lte=last, end_date__gte=first)
        summary.append([idx, user.get_full_name() or user.username, user_records.count(), user_records.filter(status="LEWAT").count(), leave_qs.count()])
    for cell in summary[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 7
    summary.column_dimensions["B"].width = 32
    for col in "CDE":
        summary.column_dimensions[col].width = 18

    output = io.BytesIO()
    wb.save(output)
    filename = f"laporan_kehadiran_{year}_{month:02d}.xlsx"
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    year, month, report_type, users, records = _report_filters(request)
    school = SchoolSettings.load()
    output = io.BytesIO()
    filename = f"laporan_kehadiran_{year}_{month:02d}.pdf"
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=28, leftMargin=28, topMargin=28, bottomMargin=28)
    styles = getSampleStyleSheet()
    centered = ParagraphStyle("Centered", parent=styles["Heading1"], alignment=TA_CENTER, fontSize=15, leading=18)
    story = [
        Paragraph(school.school_name, centered),
        Paragraph((f"Laporan Kehadiran {month:02d}/{year}" if report_type != "annual" else f"Laporan Kehadiran Tahunan {year}"), ParagraphStyle("Sub", parent=styles["Normal"], alignment=TA_CENTER, fontSize=10)),
        Paragraph(f"Dicetak oleh: {request.user.get_full_name() or request.user.username} | {timezone.localtime():%d/%m/%Y %H:%M}", ParagraphStyle("Meta", parent=styles["Normal"], alignment=TA_CENTER, fontSize=8)),
        Spacer(1, 14),
    ]
    data = [["Bil", "Tarikh", "Nama Guru", "Masuk", "Keluar", "Status", "Lewat"]]
    for idx, r in enumerate(records[:500], 1):
        target_in, _ = school.times_for_date(r.date)
        late_text = "-"
        if r.check_in and r.status == "LEWAT":
            local_in = timezone.localtime(r.check_in)
            target_dt = timezone.make_aware(datetime.combine(r.date, target_in), timezone.get_current_timezone())
            late_text = f"{max(int((local_in-target_dt).total_seconds()//60),0)} min"
        data.append([
            idx, r.date.strftime("%d/%m/%Y"), r.user.get_full_name() or r.user.username,
            timezone.localtime(r.check_in).strftime("%H:%M") if r.check_in else "-",
            timezone.localtime(r.check_out).strftime("%H:%M") if r.check_out else "-",
            r.get_status_display(), late_text,
        ])
    table = Table(data, repeatRows=1, colWidths=[35, 70, 190, 65, 65, 70, 65])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#dbeafe")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#1e293b")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("GRID", (0,0), (-1,-1), .35, colors.HexColor("#94a3b8")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
    ]))
    story.append(table)
    doc.build(story)
    output.seek(0)
    return FileResponse(output, as_attachment=True, filename=filename)


@user_passes_test(lambda u: u.is_staff)
def attendance_map(request):
    today = timezone.localdate()
    selected_date = today
    raw_date = request.GET.get("date", "").strip()
    if raw_date:
        try:
            selected_date = datetime.strptime(raw_date, "%Y-%m-%d").date()
        except ValueError:
            messages.warning(request, "Format tarikh tidak sah. Tarikh hari ini digunakan.")

    teacher_id = request.GET.get("teacher", "").strip()
    status = request.GET.get("status", "").strip().upper()
    records = Attendance.objects.filter(
        date=selected_date, check_in__isnull=False
    ).select_related("user").order_by("check_in")

    if teacher_id.isdigit():
        records = records.filter(user_id=int(teacher_id))
    if status in {"HADIR", "LEWAT"}:
        records = records.filter(status=status)

    markers = []
    for rec in records:
        name = rec.user.get_full_name() or rec.user.username
        if rec.check_in_lat is not None and rec.check_in_lng is not None:
            markers.append({
                "type": "Masuk",
                "name": name,
                "lat": rec.check_in_lat,
                "lng": rec.check_in_lng,
                "time": timezone.localtime(rec.check_in).strftime("%H:%M:%S") if rec.check_in else "—",
                "status": rec.get_status_display(),
                "accuracy": round(rec.check_in_accuracy, 1) if rec.check_in_accuracy is not None else None,
                "distance": round(rec.distance_in_m, 1) if rec.distance_in_m is not None else None,
                "device": rec.check_in_device or "Peranti tidak direkod",
                "google_url": f"https://www.google.com/maps?q={rec.check_in_lat},{rec.check_in_lng}",
            })
        if rec.check_out_lat is not None and rec.check_out_lng is not None:
            markers.append({
                "type": "Keluar",
                "name": name,
                "lat": rec.check_out_lat,
                "lng": rec.check_out_lng,
                "time": timezone.localtime(rec.check_out).strftime("%H:%M:%S") if rec.check_out else "—",
                "status": rec.get_status_display(),
                "accuracy": round(rec.check_out_accuracy, 1) if rec.check_out_accuracy is not None else None,
                "distance": round(rec.distance_out_m, 1) if rec.distance_out_m is not None else None,
                "device": rec.check_out_device or "Peranti tidak direkod",
                "google_url": f"https://www.google.com/maps?q={rec.check_out_lat},{rec.check_out_lng}",
            })

    User = get_user_model()
    teachers = User.objects.filter(is_active=True, is_staff=False).order_by("first_name", "username")
    school = SchoolSettings.load()
    return render(request, "attendance/attendance_map.html", {
        "school": school,
        "selected_date": selected_date,
        "selected_teacher": teacher_id,
        "selected_status": status,
        "teachers": teachers,
        "records": records,
        "markers": markers,
    })

@user_passes_test(lambda u: u.is_staff)
def import_teachers(request):
    result = None
    if request.method == "POST":
        form = TeacherImportForm(request.POST, request.FILES)
        if form.is_valid():
            from openpyxl import load_workbook
            wb = load_workbook(form.cleaned_data["file"], data_only=True)
            ws = wb.active
            headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}
            required = ["nama pengguna", "nama penuh", "emel", "no. staf", "jawatan"]
            missing = [h for h in required if h not in headers]
            if missing:
                messages.error(request, "Lajur tiada: " + ", ".join(missing))
            else:
                User = get_user_model()
                created = updated = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    username = str(row[headers["nama pengguna"]] or "").strip()
                    full_name = str(row[headers["nama penuh"]] or "").strip()
                    if not username or not full_name:
                        continue
                    user, is_new = User.objects.get_or_create(username=username)
                    parts = full_name.split(maxsplit=1)
                    user.first_name = parts[0]
                    user.last_name = parts[1] if len(parts) > 1 else ""
                    user.email = str(row[headers["emel"]] or "").strip()
                    if is_new:
                        user.set_password(form.cleaned_data["default_password"])
                        created += 1
                    else:
                        updated += 1
                    user.save()
                    profile, _ = TeacherProfile.objects.get_or_create(user=user)
                    profile.staff_id = str(row[headers["no. staf"]] or "").strip()
                    profile.position = str(row[headers["jawatan"]] or "").strip()
                    profile.save()
                result = {"created": created, "updated": updated}
                messages.success(request, "Import guru selesai.")
    else:
        form = TeacherImportForm()
    return render(request, "attendance/import_teachers.html", {"form": form, "result": result})


def password_recovery_request(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    form = PasswordRecoveryRequestForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        username = form.cleaned_data["username"].strip()
        User = get_user_model()
        user = User.objects.filter(username__iexact=username, is_active=True).first()
        if user:
            existing = PasswordRecoveryRequest.objects.filter(user=user, status__in=["MENUNGGU", "DILULUSKAN"]).first()
            if not existing:
                PasswordRecoveryRequest.objects.create(user=user)
                AccountActivity.objects.create(user=user, action="Minta reset kata laluan", details="Permintaan dihantar kepada pentadbir")
                from .push import send_notification
                for admin_user in User.objects.filter(is_staff=True, is_active=True):
                    send_notification(admin_user, "Permintaan reset kata laluan", f"{user.get_full_name() or user.username} meminta QR reset kata laluan.", "KESELAMATAN", "/pemulihan-kata-laluan/admin/")
        messages.success(request, "Permintaan telah dihantar. Hubungi pentadbir sekolah untuk mengimbas QR reset.")
        return redirect("password_recovery_status")
    return render(request, "attendance/password_recovery_request.html", {"form": form})


def password_recovery_status(request):
    return render(request, "attendance/password_recovery_status.html")


def password_recovery_confirm(request):
    """Fallback manual-code recovery for devices that cannot scan QR."""
    if request.user.is_authenticated:
        return redirect("dashboard")
    form = PasswordRecoveryConfirmForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        username = form.cleaned_data["username"].strip()
        User = get_user_model()
        user = User.objects.filter(username__iexact=username, is_active=True).first()
        recovery = None
        if user:
            recovery = PasswordRecoveryRequest.objects.filter(user=user, status="DILULUSKAN").order_by("-approved_at").first()
        valid = bool(recovery and recovery.expires_at and recovery.expires_at > timezone.now() and check_password(form.cleaned_data["code"], recovery.code_hash))
        if not valid:
            messages.error(request, "Kod tidak sah, telah tamat tempoh atau belum diluluskan.")
        else:
            user.set_password(form.cleaned_data["new_password1"])
            user.save()
            recovery.status = "SELESAI"
            recovery.used_at = timezone.now()
            recovery.code_display = ""
            recovery.save(update_fields=["status", "used_at", "code_display"])
            AccountActivity.objects.create(user=user, action="Reset kata laluan", details="Kata laluan dipulihkan menggunakan kod pentadbir")
            messages.success(request, "Kata laluan berjaya ditukar. Sila log masuk.")
            return redirect("login")
    return render(request, "attendance/password_recovery_confirm.html", {"form": form})


def password_recovery_qr_scan(request, pk, code):
    """Validate the short-lived QR and authorize a one-time password change."""
    if request.user.is_authenticated:
        return redirect("dashboard")
    recovery = get_object_or_404(PasswordRecoveryRequest.objects.select_related("user"), pk=pk)
    valid = (
        recovery.status == "DILULUSKAN"
        and recovery.expires_at
        and recovery.expires_at > timezone.now()
        and check_password(code, recovery.code_hash)
    )
    if not valid:
        return render(request, "attendance/password_recovery_qr_invalid.html", status=400)
    request.session["password_recovery_qr_id"] = recovery.pk
    request.session.set_expiry(15 * 60)
    return redirect("password_recovery_qr_set")


def password_recovery_qr_set(request):
    if request.user.is_authenticated:
        return redirect("dashboard")
    recovery_id = request.session.get("password_recovery_qr_id")
    recovery = PasswordRecoveryRequest.objects.select_related("user").filter(pk=recovery_id).first()
    valid = bool(recovery and recovery.status == "DILULUSKAN" and recovery.expires_at and recovery.expires_at > timezone.now())
    if not valid:
        request.session.pop("password_recovery_qr_id", None)
        return render(request, "attendance/password_recovery_qr_invalid.html", status=400)

    form = QRPasswordSetForm(request.POST or None, user=recovery.user)
    if request.method == "POST" and form.is_valid():
        recovery.user.set_password(form.cleaned_data["new_password1"])
        recovery.user.save()
        recovery.status = "SELESAI"
        recovery.used_at = timezone.now()
        recovery.code_display = ""
        recovery.save(update_fields=["status", "used_at", "code_display"])
        AccountActivity.objects.create(
            user=recovery.user,
            action="Reset kata laluan QR",
            details="Kata laluan dipulihkan melalui QR pentadbir",
        )
        request.session.pop("password_recovery_qr_id", None)
        messages.success(request, "Kata laluan berjaya ditukar. Sila log masuk.")
        return redirect("login")
    return render(request, "attendance/password_recovery_qr_set.html", {"form": form, "recovery": recovery})


@user_passes_test(lambda u: u.is_staff)
def password_recovery_admin(request):
    import qrcode
    items = list(PasswordRecoveryRequest.objects.select_related("user", "approved_by")[:100])
    now = timezone.now()
    for item in items:
        item.qr_data_uri = ""
        item.qr_url = ""
        if item.status == "DILULUSKAN" and item.code_display and item.expires_at and item.expires_at > now:
            item.qr_url = request.build_absolute_uri(
                reverse("password_recovery_qr_scan", args=[item.pk, item.code_display])
            )
            image = qrcode.make(item.qr_url)
            buffer = io.BytesIO()
            image.save(buffer, format="PNG")
            item.qr_data_uri = "data:image/png;base64," + base64.b64encode(buffer.getvalue()).decode("ascii")
    return render(request, "attendance/password_recovery_admin.html", {"items": items})


@require_POST
@user_passes_test(lambda u: u.is_staff)
def password_recovery_approve(request, pk):
    import secrets
    item = get_object_or_404(PasswordRecoveryRequest.objects.select_related("user"), pk=pk)
    code = f"{secrets.randbelow(1000000):06d}"
    item.status = "DILULUSKAN"
    item.code_hash = make_password(code)
    item.code_display = code
    item.expires_at = timezone.now() + timedelta(minutes=15)
    item.approved_at = timezone.now()
    item.approved_by = request.user
    item.used_at = None
    item.save()
    AccountActivity.objects.create(user=item.user, action="QR reset diluluskan", details=f"Diluluskan oleh {request.user.username}")
    from .push import send_notification
    send_notification(item.user, "QR reset telah dijana", "Pentadbir telah menjana QR reset kata laluan yang sah selama 15 minit.", "KESELAMATAN", "/lupa-kata-laluan/status/")
    messages.success(request, f"QR reset untuk {item.user.username} telah dijana dan sah selama 15 minit.")
    return redirect("password_recovery_admin")


@require_POST
@user_passes_test(lambda u: u.is_staff)
def password_recovery_reject(request, pk):
    item = PasswordRecoveryRequest.objects.get(pk=pk)
    item.status="DITOLAK"; item.code_display=""; item.save(update_fields=["status","code_display"])
    messages.success(request, "Permintaan ditolak.")
    return redirect("password_recovery_admin")

@user_passes_test(lambda u: u.is_staff)
def audit_log_page(request):
    from django.core.paginator import Paginator
    from django.db.models import Q, Count
    from .models import SystemAuditLog

    logs = SystemAuditLog.objects.select_related("user").all()
    q = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    severity = request.GET.get("severity", "").strip()
    user_id = request.GET.get("user", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()

    if q:
        logs = logs.filter(Q(username_snapshot__icontains=q) | Q(action__icontains=q) | Q(description__icontains=q) | Q(ip_address__icontains=q) | Q(path__icontains=q))
    if category:
        logs = logs.filter(category=category)
    if severity:
        logs = logs.filter(severity=severity)
    if user_id:
        logs = logs.filter(user_id=user_id)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)

    filtered_count = logs.count()
    today = timezone.localdate()
    stats = {
        "total": SystemAuditLog.objects.count(),
        "today": SystemAuditLog.objects.filter(created_at__date=today).count(),
        "warnings": SystemAuditLog.objects.filter(severity__in=["AMARAN", "KRITIKAL"]).count(),
        "login_failed": SystemAuditLog.objects.filter(action="Percubaan log masuk gagal").count(),
    }
    category_stats = list(SystemAuditLog.objects.values("category").annotate(total=Count("id")).order_by("-total"))
    User = get_user_model()
    users = User.objects.filter(is_active=True).order_by("first_name", "username")
    page_obj = Paginator(logs, 40).get_page(request.GET.get("page"))

    return render(request, "attendance/audit_log.html", {
        "page_obj": page_obj,
        "filtered_count": filtered_count,
        "stats": stats,
        "category_stats": category_stats,
        "users": users,
        "categories": SystemAuditLog.CATEGORY_CHOICES,
        "severities": SystemAuditLog.SEVERITY_CHOICES,
        "filters": {"q": q, "category": category, "severity": severity, "user": user_id, "date_from": date_from, "date_to": date_to},
    })


@user_passes_test(lambda u: u.is_staff)
def audit_log_export_csv(request):
    import csv
    from django.db.models import Q
    from .models import SystemAuditLog
    from .audit import write_audit

    logs = SystemAuditLog.objects.select_related("user").all()
    q = request.GET.get("q", "").strip()
    category = request.GET.get("category", "").strip()
    severity = request.GET.get("severity", "").strip()
    user_id = request.GET.get("user", "").strip()
    date_from = request.GET.get("date_from", "").strip()
    date_to = request.GET.get("date_to", "").strip()
    if q:
        logs = logs.filter(Q(username_snapshot__icontains=q) | Q(action__icontains=q) | Q(description__icontains=q) | Q(ip_address__icontains=q) | Q(path__icontains=q))
    if category: logs = logs.filter(category=category)
    if severity: logs = logs.filter(severity=severity)
    if user_id: logs = logs.filter(user_id=user_id)
    if date_from: logs = logs.filter(created_at__date__gte=date_from)
    if date_to: logs = logs.filter(created_at__date__lte=date_to)

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="audit_log_{timezone.localdate():%Y%m%d}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(["Tarikh/Masa", "Pengguna", "Kategori", "Tahap", "Tindakan", "Keterangan", "IP", "Peranti", "Kaedah", "Laluan", "Status HTTP"])
    for item in logs.iterator():
        writer.writerow([
            timezone.localtime(item.created_at).strftime("%d/%m/%Y %H:%M:%S"), item.username_snapshot,
            item.get_category_display(), item.get_severity_display(), item.action, item.description,
            item.ip_address or "", item.device, item.method, item.path, item.status_code or "",
        ])
    write_audit(request=request, category="LAPORAN", action="Eksport log audit CSV", description=f"{logs.count()} rekod dieksport.", status_code=200)
    return response


@login_required
def notification_center(request):
    category = request.GET.get("category", "").upper()
    items = AppNotification.objects.filter(user=request.user)
    if category in {"KEHADIRAN", "CUTI", "KESELAMATAN", "SISTEM"}:
        items = items.filter(category=category)
    return render(request, "attendance/notifications.html", {
        "items": items[:150],
        "selected_category": category,
        "unread_count": AppNotification.objects.filter(user=request.user, is_read=False).count(),
    })


@login_required
@require_POST
def notification_mark_read(request, pk):
    from .push import mark_notification_read
    item = get_object_or_404(AppNotification, pk=pk, user=request.user)
    mark_notification_read(item)
    return redirect(item.url or "notification_center")


@login_required
@require_POST
def notification_mark_all_read(request):
    AppNotification.objects.filter(user=request.user, is_read=False).update(is_read=True, read_at=timezone.now())
    messages.success(request, "Semua notifikasi ditandakan sebagai telah dibaca.")
    return redirect("notification_center")


@login_required
def push_public_key(request):
    from .push import get_vapid_configuration
    return JsonResponse({"publicKey": get_vapid_configuration().public_key})


@login_required
@require_POST
def push_subscribe(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        endpoint = payload["endpoint"]
        keys = payload["keys"]
        p256dh = keys["p256dh"]
        auth = keys["auth"]
    except (ValueError, KeyError, TypeError):
        return JsonResponse({"ok": False, "message": "Data langganan tidak sah."}, status=400)
    subscription, _ = PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={
            "user": request.user,
            "p256dh": p256dh,
            "auth": auth,
            "user_agent": request.META.get("HTTP_USER_AGENT", "")[:2000],
            "is_active": True,
        },
    )
    return JsonResponse({"ok": True, "message": "Push Notification telah diaktifkan.", "id": subscription.pk})


@login_required
@require_POST
def push_unsubscribe(request):
    try:
        payload = json.loads(request.body.decode("utf-8"))
        endpoint = payload.get("endpoint", "")
    except ValueError:
        endpoint = ""
    if endpoint:
        PushSubscription.objects.filter(user=request.user, endpoint=endpoint).update(is_active=False)
    return JsonResponse({"ok": True, "message": "Push Notification telah dinyahaktifkan."})
